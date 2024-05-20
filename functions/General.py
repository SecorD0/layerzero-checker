import logging

from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from data import config
from utils.db_api.database import db
from utils.db_api.models import CheckingAddress
from utils.miscellaneous.read_spreadsheet import read_spreadsheet


class General:
    @staticmethod
    def import_addresses() -> None:
        try:
            imported = 0
            for row in read_spreadsheet(path=config.ADDRESSES_FILE, sheet_name='Addresses'):
                address = row.get('address')
                if address:
                    db.insert(CheckingAddress(address=address.lower()))
                    imported += 1

            print(f'{config.LIGHTGREEN_EX}{imported}{config.RESET_ALL} addresses have been imported.')

        except BaseException as e:
            logging.exception('General.import_addresses')
            print(f"{config.RED}Failed to import addresses: {e}{config.RESET_ALL}")

    @staticmethod
    def export_addresses() -> None:
        try:
            checking_addresses = list(db.execute('SELECT * FROM checking_addresses'))
            if checking_addresses:
                spreadsheet = load_workbook(config.ADDRESSES_FILE)
                if 'Results' in spreadsheet:
                    del spreadsheet['Results']

                sheet: Worksheet = spreadsheet.create_sheet('Results')
                for column, header in enumerate(
                        ['n'] + list(db.execute('SELECT * FROM checking_addresses').keys())[1:]
                ):
                    sheet.cell(row=1, column=column + 1).value = header

                for row, address in enumerate(checking_addresses):
                    cell = sheet.cell(row=row + 2, column=1)
                    cell.number_format = '0'
                    cell.value = row + 1

                    for column, value in enumerate(address[1:]):
                        sheet.cell(row=row + 2, column=column + 2).value = value

                spreadsheet.save(config.ADDRESSES_FILE)
                print(
                    f'\nDone! The results exported to the '
                    f'{config.LIGHTGREEN_EX}addresses.xlsx{config.RESET_ALL} spreadsheet.'
                )

        except BaseException as e:
            logging.exception('General.export_addresses')
            print(f"{config.RED}Failed to export addresses: {e}{config.RESET_ALL}")
